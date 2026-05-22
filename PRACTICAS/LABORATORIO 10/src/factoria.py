# src/factoria.py
import openpyxl
from dataset import DataSetRegresion
from registro import RegistroRegresion

class FactoriaSerieTemporal:
    @staticmethod
    def leer_desde_excel(ruta_archivo: str, ph: int) -> DataSetRegresion:
        """
        Lee un archivo .xlsx y crea un DataSet de Regresión usando PH (Past History).
        No usa numpy ni pandas.
        """
        # 1. Cargar el Excel (asegúrate de tener instalado: pip install openpyxl)
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        hoja = wb.active

        # 2. Extraer los valores de la serie (Columna B / Índice 1)
        valores_serie = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if fila[1] is not None:
                valores_serie.append(float(fila[1]))

        # 3. Crear el dataset y aplicar Ventana Deslizante
        dataset = DataSetRegresion()
        # Nombres de columnas sugeridos: Lag_12, Lag_11... Lag_1
        nombres = [f"Lag_{i}" for i in range(ph, 0, -1)]
        dataset.set_cabeceras(nombres)

        for i in range(ph, len(valores_serie)):
            atributos = valores_serie[i-ph : i] # Los 'ph' valores anteriores
            objetivo = valores_serie[i]        # El valor actual que queremos predecir
            
            reg = RegistroRegresion(atributos, objetivo)
            dataset.agregar_registro(reg)

        return dataset
    
    @staticmethod
    def leer_desde_excel_diferencias(ruta_archivo: str, ph: int) -> tuple[DataSetRegresion, float]:
        """
        Lee el archivo, convierte la serie original en una serie de DIFERENCIAS (y_t = x_t - x_{t-1})
        y devuelve el DataSet de regresión junto con el último valor real (para la reconstrucción).
        """
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        hoja = wb.active

        valores_serie = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if fila[1] is not None:
                valores_serie.append(float(fila[1]))

        # Guardamos el último valor original (servirá como ancla para la primera predicción)
        ultimo_valor_real = valores_serie[-1]

        # 1. Transformamos la serie a diferencias
        valores_diff = []
        for i in range(1, len(valores_serie)):
            diferencia = valores_serie[i] - valores_serie[i-1]
            valores_diff.append(diferencia)

        # 2. Crear el dataset y aplicar Ventana Deslizante (PH) sobre las diferencias
        dataset = DataSetRegresion()
        nombres = [f"Lag_{i}" for i in range(ph, 0, -1)]
        dataset.set_cabeceras(nombres)

        for i in range(ph, len(valores_diff)):
            atributos = valores_diff[i-ph : i] 
            objetivo = valores_diff[i]
            dataset.agregar_registro(RegistroRegresion(atributos, objetivo))

        return dataset, ultimo_valor_real